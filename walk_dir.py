# -*- coding: utf-8 -*-
#遍历文件夹取文件名
import os,os.path
import re
from openpyxl import Workbook,load_workbook
import json
import datetime
import pickle

# import os


dir_name = u"E:\kuaipan\非投标任务\\2016年5月3日 micheal 要的投标记录统计\开标记录\\2014年开标汇总及分析表"
# dir_name = u"E:\\kuaipan\\非投标任务\\2016年5月3日 micheal 要的投标记录统计\\开标记录"
dir_name = u'E:\kuaipan\非投标任务\\2016年5月3日 micheal 要的投标记录统计\开标记录'
dir_name = u"E:\kuaipan\非投标任务\\2016年5月3日 micheal 要的投标记录统计\开标记录\\2013年开标汇总及分析表"

def walk_dir(dir,topdown = True):
    bid_arr = []
    for root,dirs,files in os.walk(dir,topdown):
        for f in files:
            #filter
            patten = re.compile(r'NCL-[TP]1\d-\d+')
            obj_match = patten.search(f)
            if obj_match:
                # print obj_match.group(0)
                full_name = os.path.join(root,f)
                bid = Bid()
                bid.wb = full_name
                bid.shts = find_worksheets(full_name)
                # bid_arr.append(bid)
                # bid_arr.append(pickle.dumps(bid))
                # print bid.to_json()
                bid_arr.append(bid)
    return bid_arr       

class Bid:
    def __init__(self,workbook="",worksheets=[]):
        self.wb = workbook
        self.shts = worksheets
    def to_json(self):
        return json.dumps(self,default=lambda o:o.__dict__,sort_keys=True,indent = 4)


def find_worksheets(excel_file_name):
    shts = []
    patten = re.compile(r'.xlsx')
    if patten.search(excel_file_name):
        try:
            wb = load_workbook(excel_file_name)
        except :
            print "%s load failed" %(excel_file_name)
            return None
        else:
            wb = load_workbook(excel_file_name)
            # print excel_file_name +" loaded"
            # i = 0
        return  wb.get_sheet_names()

        # for sht in wb.get_sheet_names():
        #     i += 1
        #     print type(sht)
        #     print "   %d,%s" %(i,sht)
        #     # # print wb.get_sheet_names()

# def save_to_file(data,file_name):
#     with open(file_name,'w') as f:
#         f.write(data)

file_name = os.path.dirname(os.path.abspath(__file__))+"\\" + datetime.datetime.now().strftime('%y%m%d %H-%M-%S')+'.json'
bids = walk_dir(dir_name)
with open(file_name,'w') as f:
    for bid in bids:
        f.write(bid.to_json())

print 'saved to ' +file_name